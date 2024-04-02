#pypi.org - хранилище всех библиотек,
# тут же можно поискать дл работы с эксель: xls, xlsx
# xslrd - для чтения и форматирования эксель
# pip install openpyxl

from openpyxl import load_workbook

workbook = load_workbook("tmp/example.xlsx")
sheet = workbook.active # работа с листо эксель
print(sheet.cell(row=1, column=1).value) #содержимое первря ячейки первого столбца
for x in sheet.columns:
    print(x)

